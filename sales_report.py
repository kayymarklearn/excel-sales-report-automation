"""
Sales Report Automation
------------------------
Automatically cleans raw sales data and generates
a formatted summary report with totals by product and date.

Author: Mark (github.com/kayymarklearn)
Usage: python sales_report.py
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


# ── 1. GENERATE SAMPLE DATA ──────────────────────────────────────────────────

def generate_sample_data():
    """Creates a sample raw sales Excel file to demo the automation."""
    data = {
        "Date": [
            "2024-01-05", "2024-01-12", "2024-01-20", "2024-01-28",
            "2024-02-03", "2024-02-14", "2024-02-22",
            "2024-03-01", "2024-03-15", "2024-03-29",
            "2024-04-07", "2024-04-18", "2024-04-25",
            "2024-05-02", "2024-05-19",
        ],
        "Product": [
            "Widget A", "widget a", "WIDGET A", "Widget A",
            "Widget B", "Widget B", "widget b",
            "Widget C", "Widget C", "Widget C",
            "Widget A", "Widget B", "Widget C",
            "Widget A", "Widget B",
        ],
        "Units Sold": [
            10, 15, 8, 20, 5, 12, 9, 30, 25, 18, 14, 7, 22, 11, 16
        ],
        "Unit Price": [
            50, 50, 50, 50, 80, 80, 80, 35, 35, 35, 50, 80, 35, 50, 80
        ],
        "Region": [
            "North", "South", "North", "East", "West", "South", "North",
            "East", "West", "North", "South", "East", "West", "North", "South"
        ],
        "Salesperson": [
            "Alice", "Bob", "Alice", "Carol", "Bob", "Alice", "Carol",
            "Bob", "Carol", "Alice", "Bob", "Carol", "Alice", "Bob", "Carol"
        ],
        "Notes": [
            "Q1 sale", "", "Promo discount", None,
            "", "Q1 sale", None,
            "Bulk order", "", "Promo discount",
            "", None, "Q2 promo", "", ""
        ]
    }

    df = pd.DataFrame(data)
    df.to_excel("raw_sales_data.xlsx", index=False)
    print("✅ Sample file created: raw_sales_data.xlsx")


# ── 2. CLEAN THE DATA ─────────────────────────────────────────────────────────

def clean_data(df):
    """Cleans and standardises raw sales data."""

    # Standardise product names (fix inconsistent casing)
    df["Product"] = df["Product"].str.strip().str.title()

    # Parse dates properly
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # Drop rows with missing critical fields
    df = df.dropna(subset=["Date", "Product", "Units Sold", "Unit Price"])

    # Fill empty notes
    df["Notes"] = df["Notes"].fillna("")

    # Calculate revenue
    df["Revenue"] = df["Units Sold"] * df["Unit Price"]

    # Add month label for grouping
    df["Month"] = df["Date"].dt.strftime("%B %Y")

    return df


# ── 3. GENERATE SUMMARY TABLES ───────────────────────────────────────────────

def build_summaries(df):
    """Builds summary tables by product and by month."""

    by_product = (
        df.groupby("Product")
        .agg(
            Total_Units=("Units Sold", "sum"),
            Total_Revenue=("Revenue", "sum"),
            Avg_Unit_Price=("Unit Price", "mean"),
            Num_Transactions=("Revenue", "count"),
        )
        .reset_index()
        .sort_values("Total_Revenue", ascending=False)
    )

    by_month = (
        df.groupby("Month")
        .agg(
            Total_Units=("Units Sold", "sum"),
            Total_Revenue=("Revenue", "sum"),
            Num_Transactions=("Revenue", "count"),
        )
        .reset_index()
    )

    # Sort months chronologically
    by_month["_sort"] = pd.to_datetime(by_month["Month"], format="%B %Y")
    by_month = by_month.sort_values("_sort").drop(columns="_sort")

    by_region = (
        df.groupby("Region")
        .agg(Total_Revenue=("Revenue", "sum"))
        .reset_index()
        .sort_values("Total_Revenue", ascending=False)
    )

    return by_product, by_month, by_region


# ── 4. WRITE FORMATTED EXCEL REPORT ─────────────────────────────────────────

def style_header_row(ws, row, num_cols, fill_color="1F4E79"):
    """Applies header styling to a row."""
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="FFFFFF")
    border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def style_data_rows(ws, start_row, end_row, num_cols):
    """Applies alternating row colours and borders to data rows."""
    light = PatternFill("solid", fgColor="D6E4F0")
    border_side = Side(style="thin", color="CCCCCC")
    border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = light
            cell.border = border
            cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    """Auto-fits column widths."""
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4


def write_report(df, by_product, by_month, by_region, output_file):
    """Writes a fully formatted Excel report."""

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # ── Sheet 1: Clean Data ──
        df_export = df.drop(columns=["Month"])
        df_export["Date"] = df_export["Date"].dt.strftime("%Y-%m-%d")
        df_export.to_excel(writer, sheet_name="Clean Data", index=False)

        # ── Sheet 2: By Product ──
        by_product.to_excel(writer, sheet_name="By Product", index=False)

        # ── Sheet 3: By Month ──
        by_month.to_excel(writer, sheet_name="By Month", index=False)

        # ── Sheet 4: By Region ──
        by_region.to_excel(writer, sheet_name="By Region", index=False)

    # ── Apply Styling ──
    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        num_cols = ws.max_column
        num_rows = ws.max_row

        style_header_row(ws, 1, num_cols)
        if num_rows > 1:
            style_data_rows(ws, 2, num_rows, num_cols)

        # Format revenue/currency columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in ["C", "D"] and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

        auto_width(ws)

    # ── Summary Sheet ──
    ws_summary = wb.create_sheet("📊 Summary", 0)

    generated_on = datetime.now().strftime("%d %B %Y, %H:%M")
    total_revenue = df["Revenue"].sum()
    total_units = df["Units Sold"].sum()
    total_transactions = len(df)
    top_product = by_product.iloc[0]["Product"]
    top_region = by_region.iloc[0]["Region"]

    summary_data = [
        ["SALES REPORT SUMMARY", ""],
        ["Generated on", generated_on],
        ["", ""],
        ["Total Revenue", f"${total_revenue:,.2f}"],
        ["Total Units Sold", total_units],
        ["Total Transactions", total_transactions],
        ["Top Product", top_product],
        ["Top Region", top_region],
    ]

    for row_data in summary_data:
        ws_summary.append(row_data)

    # Style the summary sheet title
    ws_summary["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws_summary["A1"].alignment = Alignment(horizontal="left")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    # Bold the labels
    for row in ws_summary.iter_rows(min_row=2):
        row[0].font = Font(bold=True)

    wb.save(output_file)
    print(f"✅ Report saved: {output_file}")


# ── 5. MAIN ───────────────────────────────────────────────────────────────────

def main():
    input_file = "raw_sales_data.xlsx"

    # Generate sample data if no input file exists
    if not os.path.exists(input_file):
        print("📄 No input file found. Generating sample data...")
        generate_sample_data()

    print("🔄 Reading and cleaning data...")
    df_raw = pd.read_excel(input_file)
    df_clean = clean_data(df_raw)

    print("📊 Building summaries...")
    by_product, by_month, by_region = build_summaries(df_clean)

    output_file = "sales_report_output.xlsx"
    print("📝 Writing formatted report...")
    write_report(df_clean, by_product, by_month, by_region, output_file)

    print("\n🎉 Done! Open sales_report_output.xlsx to see your report.")


if __name__ == "__main__":
    main()
